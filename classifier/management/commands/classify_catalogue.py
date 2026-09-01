"""
Usage:
    python manage.py classify_catalogue "/path/to/Product List.xlsx" --limit 5000
    python manage.py classify_catalogue "/path/to/Product List.xlsx" --limit 200 --with-images
    python manage.py classify_catalogue "/path/to/Product List.xlsx" --limit 200 --with-images --with-vision --with-llm
    python manage.py classify_catalogue "/path/to/Product List.xlsx" --limit 5000 --async   # dispatch to Celery workers

PRIORITY 1 FIX (header-based import): columns are now located by matching
the spreadsheet's own header row, not by fixed column index. This is
resilient to columns being reordered or inserted in a different export of
the same catalogue -- the previous version (`COL_SKU = 0`, etc.) would
silently read the wrong column if that happened. Header names are matched
case-insensitively against a list of accepted aliases per field (see
HEADER_ALIASES) so minor naming variations ("SKU" vs "Product Number")
still resolve correctly; if no alias matches, that field is imported as
empty (never guessed from position) and a warning is printed once at the
start of the import telling you which fields couldn't be located.

PRIORITY 1 FIX (multiple images): all "Image N" columns are now collected
into Product.image_urls (up to 20 in this catalogue), not just the first one.
Product.image_url still holds the first available image for convenience
(e.g. for the color-extraction layer, which only needs one representative
photo), but the full list is stored and exposed via the API/dashboard.

PRIORITY 4 (product type): if the catalogue has no column matching any of
the accepted "Product Type" aliases (this one doesn't), product_type is
imported as "" -- never invented. See engine.py and the README for how the
classifier falls back to title/description/brand/image when it's empty.

IMPORTANT (unchanged from v2/v3): classification uses ONLY title,
description, product_type, brand, and (optionally) image -- never the
spreadsheet's existing Product Category / Product Sub Category columns.
Those are imported and stored for REFERENCE/COMPARISON only. See
engine.py's module docstring for why.
"""

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from classifier.models import (
    Product, ProductClassification, ClassificationJob,
    Category, Attribute, AttributeValue, ProductAttributeValue,
)
from classifier.services.engine import classify_product
from classifier.services.image_analysis import analyze_image
from classifier.services.text_extraction import extract_brand

CHUNK_SIZE = 100
PROGRESS_EVERY = 20  # print a progress line every N products (finer than CHUNK_SIZE, so a live demo shows movement within each chunk, not just at chunk boundaries -- Priority 2)

# Accepted header names per field, matched case-insensitively. Add aliases
# here if you run this against a differently-named export of the catalogue.
HEADER_ALIASES = {
    "sku": ["product number", "sku", "model number", "item number"],
    "category": ["product category", "category"],
    "sub_category": ["product sub category", "sub category", "subcategory"],
    "product_type": ["product type", "type"],  # not present in this catalogue -- see Priority 4
    "title": ["product name", "title", "name"],
    "description": ["product description", "description"],
}
IMAGE_HEADER_PREFIX = "image"  # matches "Image 1".."Image 20"


class Command(BaseCommand):
    help = "Import a product catalogue (header-based column detection) and run resumable batch classification."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)
        parser.add_argument("--limit", type=int, default=300, help="Max rows to import (default 300 for a quick demo)")
        parser.add_argument("--with-images", action="store_true", help="Actually download and analyze the primary product image (slow)")
        parser.add_argument("--with-vision", action="store_true", help="Use the optional vision-LLM category layer if ANTHROPIC_API_KEY is set (requires --with-images)")
        parser.add_argument("--with-llm", action="store_true", help="Use the optional text-LLM layer if ANTHROPIC_API_KEY is set")
        parser.add_argument("--async", dest="run_async", action="store_true", help="Dispatch to Celery workers instead of processing in this process (requires Celery+Redis running -- see README Priority 7)")

    def handle(self, *args, **options):
        xlsx_path = options["xlsx_path"]
        limit = options["limit"]
        with_images = options["with_images"]
        with_vision = options["with_vision"]
        with_llm = options["with_llm"]
        run_async = options["run_async"]

        column_map, image_columns = self._resolve_columns(xlsx_path)
        self._report_column_mapping(column_map, image_columns)

        imported = self._import_spreadsheet(xlsx_path, limit, column_map, image_columns)
        self.stdout.write(f"Imported {imported} new products.")

        if run_async:
            self._dispatch_to_celery(with_images, with_vision, with_llm)
        else:
            if with_images:
                self.stdout.write(self.style.WARNING("--with-images enabled: fetching real product images (this is slower)."))
            self._run_classification_job_sync(with_images, with_vision, with_llm)

    # -- Column resolution (Priority 1) --------------------------------------

    def _resolve_columns(self, xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise CommandError(f"Could not read a header row from {xlsx_path}")

        headers_lower = [(str(h).strip().lower() if h else "") for h in header_row]

        column_map = {}
        for field, aliases in HEADER_ALIASES.items():
            column_map[field] = next(
                (i for i, h in enumerate(headers_lower) if h in aliases), None
            )

        image_columns = [i for i, h in enumerate(headers_lower) if h.startswith(IMAGE_HEADER_PREFIX)]
        return column_map, image_columns

    def _report_column_mapping(self, column_map, image_columns):
        missing = [field for field, idx in column_map.items() if idx is None]
        self.stdout.write("Column mapping (by header name, not position):")
        for field, idx in column_map.items():
            self.stdout.write(f"  {field:15s} -> {'column ' + str(idx) if idx is not None else '(not found -- imported as empty)'}")
        self.stdout.write(f"  {'images':15s} -> {len(image_columns)} image column(s) found")
        if missing:
            self.stdout.write(self.style.WARNING(
                f"Note: no column found for {missing} -- these fields will be empty for every row. "
                f"If your spreadsheet uses different header names, add them to HEADER_ALIASES in this command."
            ))

    # -- Import ---------------------------------------------------------------

    def _import_spreadsheet(self, xlsx_path, limit, column_map, image_columns):
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active

        existing_skus = set(Product.objects.values_list("sku", flat=True))
        to_create = []
        inserted = 0

        def get(row, field):
            idx = column_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if inserted >= limit:
                break
            try:
                sku = get(row, "sku")
                if not sku or sku in existing_skus:
                    continue
                title = get(row, "title") or ""
                all_images = [row[idx] for idx in image_columns if idx < len(row) and row[idx]]

                to_create.append(Product(
                    sku=sku,
                    title=title,
                    description=get(row, "description") or "",
                    product_type=get(row, "product_type") or "",  # "" when the catalogue has no such column (Priority 4)
                    brand=extract_brand(title),  # extracted from title text, not a source-data lookup (Q6)
                    # Stored for REFERENCE/COMPARISON only -- never fed into the classifier.
                    source_category=get(row, "category") or "",
                    source_sub_category=get(row, "sub_category") or "",
                    image_url=all_images[0] if all_images else "",
                    image_urls=all_images,
                ))
                existing_skus.add(sku)
                inserted += 1
            except Exception as exc:  # malformed row must not stop the import (Q8)
                self.stderr.write(f"  skipped malformed row {i}: {exc}")
                continue

        with transaction.atomic():
            Product.objects.bulk_create(to_create, batch_size=CHUNK_SIZE)
            new_skus = [p.sku for p in to_create]
            new_products = Product.objects.filter(sku__in=new_skus).exclude(
                id__in=ProductClassification.objects.values_list("product_id", flat=True)
            )
            ProductClassification.objects.bulk_create(
                [ProductClassification(product=p, status="pending") for p in new_products],
                batch_size=CHUNK_SIZE,
            )

        return inserted

    # -- Synchronous classification (default) ----------------------------------

    def _run_classification_job_sync(self, with_images, with_vision, with_llm):
        pending_qs = ProductClassification.objects.filter(status__in=["pending", "processing"])
        total_pending = pending_qs.count()

        if total_pending == 0:
            self.stdout.write("Nothing to process -- all products already classified.")
            return

        job = ClassificationJob.objects.create(total=total_pending, mode="sync")
        processed = done = failed = review = image_failures = 0

        categories_by_path = {c.full_path: c for c in Category.objects.all()}
        attributes_by_name = {a.name.lower(): a for a in Attribute.objects.all()}

        while True:
            chunk = list(pending_qs.select_related("product")[:CHUNK_SIZE])
            if not chunk:
                break

            for classification in chunk:
                classification.status = "processing"
                classification.save(update_fields=["status"])

                product = classification.product
                image_result = None
                if with_images and product.image_url:
                    image_result = analyze_image(product.image_url)
                    if not image_result.get("processed"):
                        image_failures += 1

                result = classify_product(
                    {
                        "title": product.title, "description": product.description,
                        "product_type": product.product_type, "brand": product.brand,
                        "image_url": product.image_url,
                    },
                    use_llm=with_llm, use_vision=with_vision, image_result=image_result,
                )  # never raises -- see services/engine.py

                self._apply_result(classification, result, job, categories_by_path, attributes_by_name)

                processed += 1
                if result["status"] == "auto_classified":
                    done += 1
                elif result["status"] == "failed":
                    failed += 1
                else:
                    review += 1

                # Live per-item progress (Priority 2): a plain print every N items so a
                # live demo doesn't look hung during a chunk -- distinct from the
                # end-of-chunk summary line below, which stays for the full breakdown.
                if processed % PROGRESS_EVERY == 0 or processed == total_pending:
                    pct = round(100 * processed / total_pending)
                    self.stdout.write(f"  Processed {processed}/{total_pending} ({pct}%)")

            msg = f"  chunk complete -- {processed}/{total_pending} total (auto: {done}, needs review: {review}, failed: {failed})"
            if with_images:
                msg += f" [image failures: {image_failures}]"
            self.stdout.write(msg)

        job.finished_at = timezone.now()
        job.done, job.failed, job.needs_review = done, failed, review
        job.save()

        self.stdout.write(self.style.SUCCESS(
            f"Job #{job.id} complete: {processed} processed "
            f"({done} auto-classified, {review} needs review, {failed} failed)."
        ))
        self._report_source_agreement()

    # -- Async classification (Priority 7/8) -----------------------------------

    def _dispatch_to_celery(self, with_images, with_vision, with_llm):
        """
        Fans out one Celery task per pending product instead of looping
        in-process. Requires a running Celery worker + broker (Redis) --
        see README "Priority 7" for setup. If Celery/Redis aren't running,
        .delay() will raise a connection error immediately, which is
        surfaced here rather than silently falling back to sync (a silent
        fallback would hide a broken setup from you).
        """
        try:
            from classifier.tasks import classify_one_product
        except ImportError as exc:
            raise CommandError(f"Celery is not installed -- run: pip install celery[redis]  ({exc})")

        pending_ids = list(
            ProductClassification.objects.filter(status__in=["pending", "processing"]).values_list("id", flat=True)
        )
        if not pending_ids:
            self.stdout.write("Nothing to process -- all products already classified.")
            return

        job = ClassificationJob.objects.create(total=len(pending_ids), mode="celery")

        try:
            for classification_id in pending_ids:
                classify_one_product.delay(classification_id, job.id, with_images, with_vision, with_llm)
        except Exception as exc:
            raise CommandError(
                f"Could not reach the Celery broker (is Redis running? see README Priority 7): {exc}"
            )

        self.stdout.write(self.style.SUCCESS(
            f"Dispatched {len(pending_ids)} classification tasks to Celery (Job #{job.id}). "
            f"Run `celery -A config worker --loglevel=info` in another terminal to process them. "
            f"Check progress with: python manage.py job_status {job.id}"
        ))

    # -- Shared result-application logic (used by both sync command and Celery task) --

    @staticmethod
    def _apply_result(classification, result, job, categories_by_path, attributes_by_name):
        classification.job = job
        classification.confidence = result["confidence"]
        classification.confidence_breakdown = result.get("confidence_breakdown", {})
        classification.reasoning = result.get("reasoning", [])
        classification.alternatives = result["alternatives"]
        classification.layers_used = result["layers_used"]
        classification.failure_reason = result["reason"] or result["error"] or ""
        classification.predicted_category = categories_by_path.get(result["predicted_path"])
        classification.status = result["status"]
        classification.save()

        ProductAttributeValue.objects.filter(classification=classification).delete()
        for attr_name, value_text in result["attributes"].items():
            attribute, _ = Attribute.objects.get_or_create(name=attr_name.capitalize())
            attributes_by_name[attr_name.lower()] = attribute
            value_obj, _ = AttributeValue.objects.get_or_create(attribute=attribute, value=value_text)
            ProductAttributeValue.objects.update_or_create(
                classification=classification, attribute=attribute,
                defaults={"value": value_obj, "confidence": classification.confidence},
            )

    def _report_source_agreement(self):
        """Informational only -- see engine.py docstring: never used for scoring."""
        classifications = ProductClassification.objects.select_related("product", "predicted_category").filter(
            status="auto_classified"
        )
        agree = disagree = 0
        for c in classifications:
            source = (c.product.source_sub_category or "").lower()
            predicted = (c.predicted_category.name if c.predicted_category else "").lower()
            if not source or not predicted:
                continue
            if predicted in source or source in predicted:
                agree += 1
            else:
                disagree += 1
        total = agree + disagree
        if total:
            self.stdout.write(
                f"\nSource-category agreement check (informational only, not used for scoring): "
                f"{agree}/{total} ({round(100 * agree / total)}%) of auto-classified products' "
                f"independently-predicted category overlaps with the spreadsheet's existing sub-category."
            )
